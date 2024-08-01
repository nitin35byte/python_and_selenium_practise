import openpyxl

workbook= openpyxl.load_workbook("login.xlsx")
sheet= workbook["Sheet1"]

row_count= sheet.max_row
column_count= sheet.max_column
print(row_count)
print(column_count)

for row in range(1,row_count+1):
    for col in range(1, column_count+1):
        cell_value=sheet.cell(row=row,column=col).value
        print(cell_value , end=" ")
    print()

sheet.cell(row= row, column=1)