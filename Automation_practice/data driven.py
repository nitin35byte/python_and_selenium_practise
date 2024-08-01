import time

from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import datetime

# Initialize the WebDriver
driver = webdriver.Chrome()
driver.get("https://www.tutorialspoint.com/selenium/practice/selenium_automation_practice.php")

# Define form element locators
name_1 = driver.find_element(By.ID, "name")
email_1 = driver.find_element(By.ID, "email")
mobile_1 = driver.find_element(By.ID, "mobile")
dob_1 = driver.find_element(By.ID, "dob")
sub_1 = driver.find_element(By.ID, "subjects")
file_input = driver.find_element(By.ID, "picture")

# Define checkbox locators (example IDs used, replace with actual IDs)
checkbox_1 = driver.find_element(By.ID, "gender")  # Example checkbox for "Automation Tester"
checkbox_2 = driver.find_element(By.XPATH, "//label[text()='Sports']")  # Example checkbox for "Manual Tester"


def read_data_from_excel(filename):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Select the active sheet (you can also use sheet names if needed)
    sheet = workbook.active

    # Get the number of rows and columns
    row_count = sheet.max_row
    column_count = sheet.max_column

    print(f"Row count: {row_count}")
    print(f"Column count: {column_count}")

    # Iterate through the rows (assuming the first row is header)
    for curr_row in range(2, row_count + 1):
        user_name = sheet.cell(row=curr_row, column=1).value
        user_email = sheet.cell(row=curr_row, column=2).value
        user_mobile = sheet.cell(row=curr_row, column=3).value
        user_dob = sheet.cell(row=curr_row, column=4).value
        user_subjects = sheet.cell(row=curr_row, column=5).value
        file_path = sheet.cell(row=curr_row, column=6).value  # Assume file path is in column 6
        profession_automation = sheet.cell(row=curr_row, column=7).value  # Assume checkbox value is in column 7
        profession_manual = sheet.cell(row=curr_row, column=8).value  # Assume checkbox value is in column 8

        # Convert user_dob to string if it is a datetime object
        if isinstance(user_dob, datetime.datetime):
            user_dob = user_dob.strftime('%Y-%m-%d')  # Format date as needed

        # Fill the form fields
        name_1.clear()
        name_1.send_keys(user_name)

        email_1.clear()
        email_1.send_keys(user_email)

        mobile_1.clear()
        mobile_1.send_keys(user_mobile)

        dob_1.clear()
        dob_1.send_keys(user_dob)

        sub_1.clear()
        sub_1.send_keys(user_subjects)
        time.sleep(5)
        # Upload the file
        file_input.send_keys(file_path)
        print("photo uploaded")

        # Handle checkboxes
        if profession_automation and not checkbox_1.is_selected():
            checkbox_1.click()
        elif not profession_automation and checkbox_1.is_selected():
            checkbox_1.click()
        time.sleep(5)
        if profession_manual and not checkbox_2.is_selected():
            checkbox_2.click()
        elif not profession_manual and checkbox_2.is_selected():
            checkbox_2.click()
        print("checkbox selected")
        # Submit the form or perform other actions
        # submit_button = driver.find_element(By.ID, "submit")
        # submit_button.click()


# Call the function with the path to your Excel file
read_data_from_excel("login.xlsx")

# Close the driver
driver.quit()
